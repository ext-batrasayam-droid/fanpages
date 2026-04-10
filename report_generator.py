import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import io

def generate_excel(data, title="YouTube Shorts Report"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Channel Report"
    _style_sheet(ws, data, title)

    # ── Sheet 2: Top Videos ───────────────────────────────────────────
    ws2 = wb.create_sheet("Top Videos")
    _top_videos_sheet(ws2, data)

    # ── Sheet 3: Stats Summary ────────────────────────────────────────
    ws3 = wb.create_sheet("Summary Stats")
    _summary_sheet(ws3, data)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_sheet(ws, data, title):
    RED = "FF0000"
    DARK_RED = "CC0000"
    WHITE = "FFFFFF"
    ALT = "FFF5F5"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # Title
    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value = f"{title}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    tc.font = Font(bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=DARK_RED)
    tc.alignment = center
    ws.row_dimensions[1].height = 32

    headers = ["#","Channel","Country","Category","Subscribers","Total Views",
               "Videos","Avg Views","Engagement %","Added By","Channel URL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = center
        c.border = _border()
    ws.row_dimensions[2].height = 22

    for i, row in enumerate(data, 1):
        r = i + 2
        fill = PatternFill("solid", fgColor=ALT if i % 2 == 0 else WHITE)
        vals = [
            i, row.get("channel_name",""), row.get("country","N/A"),
            row.get("category",""), row.get("subscribers",0),
            row.get("total_views",0), row.get("video_count",0),
            row.get("avg_views_per_video",0), row.get("engagement_rate",0),
            row.get("added_by",""), row.get("url","")
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.border = _border()
            c.alignment = left if col in (2, 11) else center
            if col == 9:  # engagement color
                try:
                    v = float(val)
                    c.font = Font(color="1a7a1a" if v >= 10 else ("b35a00" if v >= 3 else "cc0000"), bold=True)
                except:
                    pass

    widths = [4,28,10,14,16,16,10,14,16,14,42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:K{len(data)+2}"

def _top_videos_sheet(ws, data):
    RED, WHITE = "FF0000", "FFFFFF"
    center = Alignment(horizontal="center", vertical="center")
    headers = ["Channel","Video Title","Views","Likes","Comments","Published","URL"]
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value = "Top Videos Per Channel"
    tc.font = Font(bold=True, size=13, color=WHITE)
    tc.fill = PatternFill("solid", fgColor="CC0000")
    tc.alignment = center
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = center
        c.border = _border()
    r = 3
    import json
    for row in data:
        try:
            videos = json.loads(row.get("top_videos_json", "[]")) if isinstance(row.get("top_videos_json"), str) else row.get("top_videos", [])
        except:
            videos = []
        for v in videos:
            vals = [row.get("channel_name",""), v.get("title",""), v.get("views",0),
                    v.get("likes",0), v.get("comments",0), v.get("published",""), v.get("url","")]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = _border()
                c.alignment = Alignment(horizontal="left" if col in (1,2,7) else "center", vertical="center")
            r += 1
    widths = [28,45,14,12,12,12,42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _summary_sheet(ws, data):
    if not data:
        return
    subs = [r.get("subscribers",0) for r in data]
    views = [r.get("total_views",0) for r in data]
    eng = [r.get("engagement_rate",0) for r in data]
    stats = [
        ("Total Channels", len(data)),
        ("Total Subscribers", sum(subs)),
        ("Total Views", sum(views)),
        ("Avg Subscribers", int(sum(subs)/len(subs))),
        ("Avg Views", int(sum(views)/len(views))),
        ("Avg Engagement %", round(sum(eng)/len(eng), 2)),
        ("Max Subscribers", max(subs)),
        ("Min Subscribers", min(subs)),
        ("Max Engagement %", max(eng)),
    ]
    ws["A1"].value = "Summary Statistics"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="CC0000")
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")
    for i, (label, val) in enumerate(stats, 2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=1).border = _border()
        ws.cell(row=i, column=2).border = _border()
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
